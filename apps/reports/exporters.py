from datetime import date
from io import BytesIO

from django.db.models import F, Sum
from openpyxl import Workbook
from openpyxl.styles import Font

from apps.accounts.models import Employee
from apps.orders.models import OrderItem
from apps.reports.services import get_period_report
from apps.salary.services import calculate_employee_salary


def export_period_excel(date_from: date, date_to: date) -> BytesIO:
    report = get_period_report(date_from, date_to)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Выручка"
    ws1.append(["Дата", "Выручка", "Расходы", "Прибыль"])
    for row in report["daily_breakdown"]:
        ws1.append(
            [
                row["date"].strftime("%d.%m.%Y"),
                float(row["total_revenue"] or 0),
                float(row["total_expenses"] or 0),
                float(row["net_profit"] or 0),
            ]
        )

    ws2 = wb.create_sheet("Расходы")
    ws2.append(["Категория", "Сумма"])
    for row in report["expenses_by_category"]:
        ws2.append([row["category__name"], float(row["total"])])

    ws3 = wb.create_sheet("Топ блюд")
    ws3.append(["Блюдо", "Кол-во", "Выручка"])
    top = (
        OrderItem.objects.filter(order__paid_at__date__range=(date_from, date_to))
        .values("menu_item__name")
        .annotate(qty=Sum("quantity"), revenue=Sum(F("price") * F("quantity")))
        .order_by("-qty")[:50]
    )
    for row in top:
        ws3.append([row["menu_item__name"], row["qty"], float(row["revenue"] or 0)])

    ws4 = wb.create_sheet("Зарплата")
    ws4.append(["Сотрудник", "Отработано", "Начислено", "К выплате"])
    header_font = Font(bold=True)
    for cell in ws4[1]:
        cell.font = header_font
    for emp in Employee.objects.filter(is_active=True):
        calc = calculate_employee_salary(emp, date_from, date_to)
        ws4.append(
            [
                str(emp),
                calc["worked_days"],
                float(calc["accrued"]),
                float(calc["net_to_pay"]),
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
